from flask import Flask, request, jsonify, render_template, session, redirect, send_file
from flask_cors import CORS
import psycopg2
import os
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
from threading import Lock

excel_lock = Lock()


file_path = "Performance_Projetos.xlsx"


if not os.path.isfile(file_path):
    with excel_lock:
        wb = Workbook()   # ✅ CRIA novo arquivo
        ws = wb.active

        ws.append([
            "ID",
            "Ideia",
            "Data",
            "Nome",
            "Matricula",
            "Supervisor",
            "Area",
            "Observacao",
            "Local_Ocorrencia",
            "Saving (R$)",

            "Qualidade",
            "Seguranca",
            "KPI",
            "Saving Score",
            "Norma",

            "Mudança Processo",
            "Mudança Material",
            "Tempo",
            "Treinamento",
            "Logistica",
            "Fornecedor",

            "Total Beneficio",
            "Total Esforco",
            "Prioridade",
            "Status"
        ])

        wb.save(file_path)   # ✅ salva o arquivo criado

load_dotenv()  

app = Flask(__name__, template_folder="templates", static_folder="static")

DATABASE_URL = os.getenv("DATABASE_URL")

app.config["SESSION_COOKIE_SAMESITE"] = "None"
app.config["SESSION_COOKIE_SECURE"] = True
app.secret_key = "incubadora-nissan"

CORS(app)

SENHA_PERFORMANCE = "123"

def conectar_db():
    return psycopg2.connect(DATABASE_URL, sslmode="require")

def sincronizar_excel():

    conn = conectar_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            id,
            ideia,
            data_criacao,
            nome,
            matricula,
            supervisor,
            area,
            observacao,
            local_ocorrencia,

            saving_real,

            qualidade,
            seguranca,
            kpi,
            saving_beneficio,
            norma,

            mudanca_processo,
            mudanca_material,
            tempo,
            treinamento,
            logistica,
            fornecedor,

            total_beneficio,
            total_esforco,
            prioridade,
            status

        FROM ideias
        ORDER BY id
    """)

    dados = cur.fetchall()

    cur.close()
    conn.close()

    with excel_lock:

        wb = load_workbook(file_path)
        ws = wb.active

        # Apaga todas as linhas de dados
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)

        # Recria o Excel usando o banco
        for r in dados:

            ws.append([
                r[0],   # ID
                r[1],   # Titulo
                str(r[2]) if r[2] else "",  # Data
                r[3],   # Nome
                r[4],   # Matricula
                r[5],   # Gestor
                r[6],   # Area
                r[7],   # Descricao
                r[8],   # Antes/Depois

                r[9],  # Saving (R$)

                r[10],  # Qualidade
                r[11],  # Seguranca
                r[12],  # KPI
                r[13],  # Saving Score
                r[14],  # Norma

                r[15],  # Mudança Processo
                r[16],  # Mudança Material
                r[17],  # Tempo
                r[18],  # Treinamento
                r[19],  # Logistica
                r[20],  # Fornecedor

                float(r[21] or 0),  # Total Beneficio
                float(r[22] or 0),  # Total Esforco
                r[23],  # Prioridade
                r[24]   # Status
            ])

        wb.save(file_path)


@app.route("/")
def index():
    if "usuario_id" not in session:
        return redirect("/login")

    return render_template(
        "index.html",
        nome=session["nome"],
        matricula=session["matricula"]
    )


@app.route("/login")
def login():
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


@app.route("/formulario")
def formulario():
    if "usuario_id" not in session:
        return redirect("/login")

    return render_template(
        "formulario.html",
        nome=session["nome"],
    matricula=session["matricula"]
)


@app.route("/performance")
def performance():
    return render_template("performance.html")


@app.route("/baixar-excel")
def baixar_excel():
    return send_file("Performance_Projetos.xlsx", as_attachment=True)

@app.route("/api/dashboard")
def api_dashboard():

    conn = conectar_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            id,
            ideia,
            data_criacao,
            nome,
            matricula,
            supervisor,
            area,
            observacao,
            local_ocorrencia,

            saving_real,

            qualidade,
            seguranca,
            kpi,
            saving_beneficio,
            norma,

            mudanca_processo,
            mudanca_material,
            tempo,
            treinamento,
            logistica,
            fornecedor,

            total_beneficio,
            total_esforco,
            prioridade,
            status

        FROM ideias
        ORDER BY id
    """)

    dados = cur.fetchall()

    cur.close()
    conn.close()

    return jsonify([
        {
            "id": r[0],
            "ideia": r[1],
            "data_criacao": str(r[2]),
            "nome": r[3],
            "matricula": r[4],
            "supervisor": r[5],
            "area": r[6],
            "observacao": r[7],
            "local_ocorrencia": r[8],

            "saving_real": r[9],

            "qualidade": r[10],
            "seguranca": r[11],
            "kpi": r[12],
            "saving_beneficio": r[13],
            "norma": r[14],

            "mudanca_processo": r[15],
            "mudanca_material": r[16],
            "tempo": r[17],
            "treinamento": r[18],
            "logistica": r[19],
            "fornecedor": r[20],

            "total_beneficio": r[21],
            "total_esforco": r[22],
            "prioridade": r[23],
            "status": r[24]
        }
        for r in dados
    ])

@app.route("/ranking")
def ranking():
    return render_template("ranking.html")

@app.route("/api/ranking")
def api_ranking():

    conn = conectar_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT nome, area, ideia, prioridade
        FROM ideias
        WHERE status IN ('Aplicada', 'Concluída')
        ORDER BY prioridade ASC
    """)


    dados = cur.fetchall()

    cur.close()
    conn.close()

    return jsonify([
    {
        "nome": r[0],
        "area": r[1],
        "ideia": r[2],
        "prioridade": r[3]
    }
    for r in dados
])

@app.route("/contato")
def contato():
    return render_template("contato.html")


@app.route("/api/login", methods=["POST"])
def api_login():

    d = request.get_json()

    if not d:
        return jsonify({"erro": "JSON inválido"}), 400

    nome = d.get("nome", "").strip()
    matricula = d.get("matricula", "").strip()

    if not nome or not matricula:
        return jsonify({"erro": "Nome e matrícula são obrigatórios"}), 400

    conn = conectar_db()
    cur = conn.cursor()

    # procura usuário pela matrícula
    cur.execute("""
        SELECT id, nome
        FROM usuarios
        WHERE matricula = %s
    """, (matricula,))

    user = cur.fetchone()

    primeiro_acesso = False

    # não existe -> cria automaticamente
    if not user:

        cur.execute("""
            INSERT INTO usuarios (
                nome,
                matricula,
                criado_em,
                ultimo_acesso
            )
            VALUES (%s, %s, NOW(), NOW())
            RETURNING id
        """, (nome, matricula))

        user_id = cur.fetchone()[0]

        conn.commit()

        session["usuario_id"] = user_id
        session["nome"] = nome
        session["matricula"] = matricula

        primeiro_acesso = True

    else:

        cur.execute("""
            UPDATE usuarios
            SET ultimo_acesso = NOW()
            WHERE matricula = %s
        """, (matricula,))

        conn.commit()

        session["usuario_id"] = user[0]
        session["nome"] = user[1]
        session["matricula"] = matricula

    cur.close()
    conn.close()

    return jsonify({
        "ok": True,
        "primeiro_acesso": primeiro_acesso
    })


@app.route("/api/ideias", methods=["POST"])
def enviar_ideia():

    if "usuario_id" not in session:
        return jsonify({"erro": "Não autorizado"}), 401

    nome = request.form.get("nome")
    matricula = request.form.get("matricula")
    area = request.form.get("area")
    supervisor = request.form.get("supervisor")

    observacao = request.form.get("observacao")
    ideia = request.form.get("ideia")

    local_ocorrencia = request.form.get("local_ocorrencia")
    beneficio = request.form.get("beneficio")

    frequencia = request.form.get("frequencia")
    apoio = request.form.get("apoio")

    arquivo = request.files.get("evidencia")

    caminho_arquivo = None

    if arquivo and arquivo.filename:

        os.makedirs("uploads", exist_ok=True)

        caminho_arquivo = os.path.join(
            "uploads",
            arquivo.filename
        )

        arquivo.save(caminho_arquivo)

    conn = conectar_db()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO ideias (
            usuario_id,
            nome,
            matricula,
            area,
            supervisor,
            observacao,
            ideia,
            local_ocorrencia,
            beneficio,
            frequencia,
            apoio,
            evidencia,
            status
        )
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        RETURNING id
    """, (
        session["usuario_id"],
        nome,
        matricula,
        area,
        supervisor,
        observacao,
        ideia,
        local_ocorrencia,
        beneficio,
        frequencia,
        apoio,
        caminho_arquivo,
        "Validação"
    ))

    novo_id = cur.fetchone()[0]

    conn.commit()

    cur.close()
    conn.close()

    return jsonify({
        "ok": True,
        "protocolo": f"2026_{novo_id:04d}"
    })



@app.route("/api/performance/auth", methods=["POST"])
def auth_performance():
    d = request.get_json()
    if not d:
        return jsonify({"erro": "JSON inválido"}), 400

    if d.get("senha") == SENHA_PERFORMANCE:
        session["performance_autorizado"] = True
        return jsonify({"ok": True})

    return jsonify({"erro": "Senha incorreta"}), 401


def autorizado():
    return True


@app.before_request
def proteger():
    if request.path.startswith("/api/performance") and not request.path.endswith("/auth"):
        if not autorizado():
            return jsonify({"erro": "Não autorizado"}), 401


@app.route("/api/performance/ideias")
def listar_ideias():
    conn = conectar_db()
    cur = conn.cursor()

    cur.execute("""
       SELECT
        id,
        data_criacao,
        nome,
        matricula,
        supervisor,
        area,

        observacao,
        ideia,
        local_ocorrencia,
        beneficio,
        frequencia,
        apoio,

        status,

        qualidade,
        seguranca,
        kpi,
        saving_beneficio,
        saving_real,
        norma,

        mudanca_processo,
        mudanca_material,
        tempo,
        treinamento,
        logistica,
        fornecedor,

        prioridade

    FROM ideias
    ORDER BY prioridade NULLS LAST, data_criacao DESC
    """)

    dados = cur.fetchall()

    cur.close()
    conn.close()

    return jsonify([
        {
            "id": r[0],
            "data_criacao": str(r[1]),

            "nome": r[2],
            "matricula": r[3],
            "supervisor": r[4],
            "area": r[5],

            "observacao": r[6],
            "ideia": r[7],
            "local_ocorrencia": r[8],
            "beneficio": r[9],
            "frequencia": r[10],
            "apoio": r[11],

            "status": r[12],

            "qualidade": r[13],
            "seguranca": r[14],
            "kpi": r[15],
            "saving_beneficio": r[16],
            "saving_real": r[17],
            "norma": r[18],

            "mudanca_processo": r[19],
            "mudanca_material": r[20],
            "tempo": r[21],
            "treinamento": r[22],
            "logistica": r[23],
            "fornecedor": r[24],

            "prioridade": r[25]
        }
    for r in dados
    ])


@app.route("/api/performance/deletar/<int:id>", methods=["DELETE"])
def deletar_ideia(id):

    conn = conectar_db()
    cur = conn.cursor()

    cur.execute(
        "DELETE FROM ideias WHERE id = %s",
        (id,)
    )

    conn.commit()

    cur.close()
    conn.close()

    sincronizar_excel()

    return jsonify({"ok": True})

@app.route("/api/performance/avaliar", methods=["POST"])
def avaliar_ideia():
    d = request.get_json()
    id_recebido = int(d["id"])

    conn = conectar_db()
    cur = conn.cursor()

    cur.execute("""
    UPDATE ideias SET

        status = %s,

        qualidade = %s,
        seguranca = %s,
        kpi = %s,
        saving_beneficio = %s,
        saving_real = %s,
        norma = %s,

        mudanca_processo = %s,
        mudanca_material = %s,
        tempo = %s,
        treinamento = %s,
        logistica = %s,
        fornecedor = %s,

        total_beneficio = %s,
        total_esforco = %s,
        prioridade = %s

        WHERE id = %s
        """, (

        d["status"],

        d["qualidade"],
        d["seguranca"],
        d["kpi"],
        d["saving_beneficio"],
        d["saving_real"],
        d["norma"],

        d["mudanca_processo"],
        d["mudanca_material"],
        d["tempo"],
        d["treinamento"],
        d["logistica"],
        d["fornecedor"],

        d["total_beneficio"],
        d["total_esforco"],
        d["prioridade"],

        d["id"]
    ))

    conn.commit()

    # ✅ CORREÇÃO
    cur.execute("""
    SELECT ideia,
        nome,
        matricula,
        supervisor,
        area,
        observacao,
        local_ocorrencia,
        data_criacao
    FROM ideias
    WHERE id = %s
    """, (id_recebido,))
    info = cur.fetchone()

    if not info:
        cur.close()
        conn.close()
        return jsonify({"erro": "Ideia não encontrada"}), 404

    with excel_lock:
        try:

            wb = load_workbook(file_path)
            ws = wb.active

            def normalizar(valor):
                try:
                    return int(float(valor))
                except:
                    return None

            linha_existente = None

            for row in ws.iter_rows(min_row=2):
                if normalizar(row[0].value) == id_recebido:
                    linha_existente = row
                    break

            if linha_existente:

                linha_existente[0].value = id_recebido
                linha_existente[1].value = info[0]

                data = str(info[7]) if len(info) > 7 else ""
                linha_existente[2].value = data

                linha_existente[3].value = info[1]
                linha_existente[4].value = info[2]
                linha_existente[5].value = info[3]
                linha_existente[6].value = info[4]
                linha_existente[7].value = info[5]
                linha_existente[8].value = info[6]

                linha_existente[9].value = d["saving_real"]

                linha_existente[10].value = d["qualidade"]
                linha_existente[11].value = d["seguranca"]
                linha_existente[12].value = d["kpi"]
                linha_existente[13].value = d["saving_beneficio"]
                linha_existente[14].value = d["norma"]

                linha_existente[15].value = d["mudanca_processo"]
                linha_existente[16].value = d["mudanca_material"]
                linha_existente[17].value = d["tempo"]
                linha_existente[18].value = d["treinamento"]
                linha_existente[19].value = d["logistica"]
                linha_existente[20].value = d["fornecedor"]

                linha_existente[21].value = d["total_beneficio"]
                linha_existente[22].value = d["total_esforco"]
                linha_existente[23].value = d["prioridade"]
                linha_existente[24].value = d["status"]

            else:

                ws.append([
                    id_recebido,
                    info[0],
                    str(info[7]),
                    info[1],
                    info[2],
                    info[3],
                    info[4],
                    info[5],
                    info[6],
                    
                    d["saving_real"],

                    d["qualidade"],
                    d["seguranca"],
                    d["kpi"],
                    d["saving_beneficio"],
                    d["norma"],

                    d["mudanca_processo"],
                    d["mudanca_material"],
                    d["tempo"],
                    d["treinamento"],
                    d["logistica"],
                    d["fornecedor"],

                    d["total_beneficio"],
                    d["total_esforco"],
                    d["prioridade"],
                    d["status"]
                ])

            wb.save(file_path)

            cur.close()
            conn.close()

            return jsonify({"ok": True})

        except Exception as e:
            print("ERRO GERAL:", e)

            cur.close()
            conn.close()

            return jsonify({"erro": str(e)}), 500

@app.route("/api/minhas-ideias")
def minhas_ideias():

    if "usuario_id" not in session:
        return jsonify({"erro": "Não autorizado"}), 401

    conn = conectar_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            id,
            ideia,
            observacao,
            status,
            data_criacao
        FROM ideias
        WHERE usuario_id = %s
        ORDER BY
            CASE
                WHEN status = 'Implementação' THEN 1
                WHEN status = 'Aplicada' THEN 2
                WHEN status = 'Concluída' THEN 3
                WHEN status = 'Projeto' THEN 4
                WHEN status = 'Execução Rápida' THEN 5
                WHEN status = 'Incubadora' THEN 6
                WHEN status = 'Validação' THEN 7
                WHEN status = 'Não Priorizada' THEN 8
                WHEN status = 'Não Viável' THEN 9
                ELSE 10
            END,
            data_criacao DESC
    """, (session["usuario_id"],))

    dados = cur.fetchall()

    cur.close()
    conn.close()

    return jsonify([
        {
            "id": r[0],
            "ideia": r[1],
            "observacao": r[2],
            "status": r[3] if r[3] else "Validação"
        }
        for r in dados
    ])


if __name__ == "__main__":
    print("SINCRONIZANDO EXCEL...")
    sincronizar_excel()
    print("EXCEL SINCRONIZADO!")

    app.run(debug=True)